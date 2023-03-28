from collections import OrderedDict
import json
from openpyxl import load_workbook, load_workbook
wb = load_workbook('python/data.xlsx')
# save numpy array as npy file

lws = wb.sheetnames
# print(lws)

# ws = wb.active

dictionary = []
sentence = []
sentences = []


def transform(x):
    ws = wb[lws[x]]
    row = ws.max_row
    column = ws.max_column
    print(row, column)
    for ro in range(2, 4):
        for i in range(1, 20):
            if (ws.cell(row=ro, column=i).value != ""):
                data = {
                    "pw": ws.cell(row=ro, column=i).value,
                }
            sentence.append(data)
        sentences.append(sentence)
    dictionary.append(sentences)
    with open('python/data.json', 'w', encoding='utf-8') as f:
        json.dump(dictionary, f, ensure_ascii=False, indent=4)


list = range(len(lws))

print(lws)
transform(0)
print(sentence)
