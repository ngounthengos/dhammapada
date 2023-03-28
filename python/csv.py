from collections import OrderedDict
from numpy import save
from numpy import asarray
import json
from openpyxl import load_workbook, load_workbook
wb = load_workbook('dict/glossary.xlsx')
# save numpy array as npy file

lws = wb.sheetnames
# print(lws)

# ws = wb.active

groups = ["(បុំ.)", "(ឥ.)", "(នបុំ.)", "(បុំ. នបុំ.)", "(គុ.)", "(បុ.និ.)", "(អ.ន.)", "(ធា.)", "(និ.)", "(បុំ., នបុំ.)", "(កិ.)", "(អ.)", "(វិ.ស.អ.)", "(គុ. ន.)", "(អ.ស.)", "(អ.ត.)", "(អ.ន. បុំ.)", "(កា.ស.)", "(បុំ.ឥ.)", "(បុំ.ឬនបុំ.)", "(នុំ.)", "(បុំ. បុំ.)", "(បុំ. ន-បុំ.)", "(គុ.,ន.)", "(នប.)", "(អា.ឥ.)", "(.បុំ.)", "(បុំ.គុ.)", "(បុំ. គុ.)", "(គុ. ឬន.)", "(បុំ.នបុំ.)", "(បុំ., គុ.)", "(កិ.វិ.)", "(គុ.,នបុំ.)", "(គុ., បុំ.)", "(ន.គុ.)", "(កិ.ឬន.)", "(បុំ.,ឥ.)", "(គុ.ន.)", "(អ.​ន.)", "(នបុំ., បុំ.)", "(ធាតុ.)", "(បុ.ស.)", "(ន. ឬ​គុ.)", "(នបុំ​.)", "(គុ.ឬ​កិ.)", "(ន.តិល.)", "(នបុំ.ឬ​គុ.)", "(ន.)", "(បុំ.,នបុំ.)", "(អស.ន.)", "(បូ.សំ.)", "(ប.)", "(និ.,គុ.)", "(ឥ. នបុំ.)", "(គ.)", "(ប.សំ.)", "(បុំ.,​នបុំ.)", "(បុំ. ឬ​គុ.)", "(គុ. ឬ​កិ.)", "(អា.និ.)", "(បុំ. ឬ​ឥ.)", "(អា.)", "(ឥ. បុំ.)", "(គុ.ឬ​បុំ.)", "(បុំ.ឬ​នបុំ.)", "(នបុំ., ឥ.)", "(បុំ. ឬ​នបុំ.)", "(វិ.ស.និ.)", "(គុ.ឬ​ន.)", "(អា.ន.)", "(នបុំ.ឬឥ.)", "(បុំ, នបុំ.)", "(បុំ.ឬគុ.)", "(កិ.គុ.)", "(បសំ.)", "(ប.ស.)", "(គុ. បុំ.)", "(បុំ. នបុំ គុ.)", "(បូ.ស.)", "(នបុំ.,គុ.)", "(អា.និ..)", "(អ., គុ.)", "(អ.ស.ន.)", "(កិ.វិ.ឬគុ.)", "(សំ.ស.)", "(គុ.ឬន.)", "(នឥ.)", "(ធាំ.)",
          "(ន. ឬ​និ.)", "(អ.ឥ.)", "(សព្វ.)", "(កិ.ន.)", "(ប,.)", "(ប,សំ.)", "(គុ. កិ.)", "(នបុំ.ឥ.)", "(នបុំ.គុ.)", "(ឝុ.)", "(នបុំ.,ឥ.)", "(គុ. ឬ​បុំ.)", "(នបុំ., និ.)", "(វិស.អ.)", "(ន.ឬ​គុ.)", "(ធា​.)", "(គុ.ឬ កិ.)", "(កិ.​ឬ​គុ.)", "(អា.ន.គុ.)", "(កិ. ឬ​គុ.)", "(នបុំ.បុំ.)", "(នបុំ.,បុំ.)", "(ឥ​.)", "(អ.កិ.)", "(ស.ន.)", "(បុំ. ឥ. នបុំ.)", "(ឥ. ប.សំ.)", "(បុំ,អ.ន.)", "(គុ.​ឬ​បុំ.)", "(កិ.ឬ​គុ.)", "(គុ.ឬ​ កិ.)", "(កិ.ឬគុ.)", "(គុ. ឬ​កិ. វិ.)", "(កិ.​គុ.)", "(ន.ឬ​កិ.គុ.)", "(គុ.ឬ​នបុំ.)", "(បុំ.នបុំឬ​គុ.)", "(ឧ.)", "(គិ.)", "(នបុំ. ឬ​ឥ.)", "(កិ.វិ. ឬ​និ.)", "(គុ.​នបុំ.)", "(នបុំ.ឬគុ.)", "(បុំ.ឬ​គុ.)", "(នបុំ, ឬឥ.)", "(លិ.វិ.)", "(គុ. នបុំ.)", "(គុណ.)", "(បុំ. អ.ន.)", "(អ. ឬកិ.វិ.)", "(វិ.ស.-និ.)", "(ឥ.,គុ.)", "(អា.និ,.)", "(ឥ. ឬនបុំ.)", "(រិ.ស.-និ.)", "(កំ.)", "(នបុំ,បុំ.)", "(បុំ.ឬ​ឥ.)", "(ឥ. ឬ​គុ.)", "(កិគុ.)", "(បុំ., នបុំ.ឬ​គុ.)", "(គុ. ឬ បុំ.)", "(នបុំ.​ឥ.)", "(នបុំ. បុំ.)", "(បុំ. ឬគុ.)", "(អ. ន.)", "(ឧប.)", "(នបុំ., គុ.)", "(បូ.សំ. គុ.)", "(បុំ. ឥ.)", "(នបុំ. ឥ.)", "(គុ.ឬកិ.)", "(គុំ.)", "(កិ.,គុ.)", "(គុ.,កិ.)", "(សំ.)", "(អ.និ.)", "(កា.និ.)", "(គុ.បុំ.)", "(កិ.គុ. នបុំ.)", "(ធា. ធា.)", "(គុ.ឥ.)", "(ឧ.ស.)", "(គុ..)"]

dictionary = []
logs = []
subs = OrderedDict([("​", ""), (' ', ""), ('។', " ។"), (' ,', ','), (' ;', ';'), (' +', "+"), ('+ ', "+"), ('+', " + "),
                   (' -', "-"), ('- ', "-"), ('-', " - "), (' >', ">"), ('> ', ">"), ('>', " > "), ("  ", ""), ("  ", "")])
# rep = {"condition1": "", "condition2": "text"}
rep = {"​": "", ' ': "", ' ។': "។", '។': " ។", ' ,': ',', ' ;': ';', ' +': "+", '+ ': "+", '+': " + ",
       ' -': "-", '- ': "-", '-': " - ", ' >': ">", '> ': ">", '>': " > ", ' <': "<", '< ': "<", '<': " < ", "  ": "", "  ": ""}
gros = {"ធា.": "ធាតុ", "ធា. ធា.": "ធាតុ", "ធា​.": "ធាតុ", "ធាតុ.": "ធាតុ"}


def transform(x):
    ws = wb[lws[x]]
    row = ws.max_row
    column = ws.max_column
    print(row, column)
    for col in ws.iter_cols(1, column):
        if row >= 3:
            if col[2].value != None:
                for g in groups:
                    if g in col[2].value:  # type: ignore
                        col[1].value = g
                if col[1].value == None:
                    col[1].value = "(មិន)"
                    # val = col[2].value
                    # if "(" in val[0:10]:
                    #     print(col[2].value)
                    #     logs.append(col[2].value)
                    # else:
                    #     col[1].value = "(មិន)"

    for col in ws.iter_cols(1, column):
        if row >= 6:
            if col[5].value != None:
                for g in groups:
                    if g in str(col[5].value):
                        col[4].value = g
                if col[4].value == None:
                    col[4].value = "(មិន)"
                    # val = col[5].value
                    # if "(" in val[0:10]:
                    #     print(col[5].value)
                    #     logs.append(col[5].value)
                    # else:
                    #     col[4].value = "(មិន)"

    for col in ws.iter_cols(1, column):
        if row >= 9:
            if col[8].value != None:
                for g in groups:
                    if g in str(col[8].value):
                        col[7].value = g
                if col[7].value == None:
                    col[7].value = "(មិន)"
                    # val = col[8].value
                    # if "(" in val[0:10]:
                    #     print(col[8].value)
                    #     logs.append(col[8].value)
                    # else:
                    #     col[7].value = "(មិន)"

    for col in ws.iter_cols(1, column):
        if row >= 12:
            if col[11].value != None:
                for g in groups:
                    if g in str(col[11].value):
                        col[10].value = g
                if col[10].value == None:
                    col[10].value = "(មិន)"
                    val = col[11].value
                    # if "(" in val[0:10]:
                    #     print(col[11].value)
                    #     logs.append(col[11].value)
                    # else:
                    #     col[10].value = "(មិន)"

    for i in range(1, column+1):
        d1 = replace_all(str(ws.cell(row=3, column=i).value), subs)
        d2 = replace_all(str(ws.cell(row=6, column=i).value), subs)
        d3 = replace_all(str(ws.cell(row=9, column=i).value), subs)
        d4 = replace_all(str(ws.cell(row=12, column=i).value), subs)
        g1 = str(ws.cell(row=2, column=i).value)
        g2 = str(ws.cell(row=5, column=i).value)
        g3 = str(ws.cell(row=8, column=i).value)
        g4 = str(ws.cell(row=11, column=i).value)
        if ws.cell(row=9, column=i).value != None:
            ws.cell(row=4, column=i).value = d1
            ws.cell(row=7, column=i).value = d2
            ws.cell(row=10, column=i).value = d3
            ws.cell(row=13, column=i).value = d4
            data = {
                "w": ws.cell(row=1, column=i).value,
                "d": [
                    {
                        "g": replace_all(g1, gros),
                        "e": str(ws.cell(row=4, column=i).value).replace(g1, "").strip(),
                    },
                    {
                        "g": replace_all(g2, gros),
                        "e": str(ws.cell(row=7, column=i).value).replace(g2, "").strip(),
                    },
                    {
                        "g": replace_all(g3, gros),
                        "e": str(ws.cell(row=10, column=i).value).replace(g3, "").strip(),
                    },
                    {
                        "g": replace_all(g4, gros),
                        "e": str(ws.cell(row=13, column=i).value).replace(g3, "").strip(),
                    }
                ]
            }
            dictionary.append(data)
        if ws.cell(row=9, column=i).value != None:
            ws.cell(row=4, column=i).value = d1
            ws.cell(row=7, column=i).value = d2
            ws.cell(row=10, column=i).value = d3
            data = {
                "w": ws.cell(row=1, column=i).value,
                "d": [
                    {
                        "g": replace_all(g1, gros),
                        "e": str(ws.cell(row=4, column=i).value).replace(g1, "").strip(),
                    },
                    {
                        "g": replace_all(g2, gros),
                        "e": str(ws.cell(row=7, column=i).value).replace(g2, "").strip(),
                    },
                    {
                        "g": replace_all(g3, gros),
                        "e": str(ws.cell(row=10, column=i).value).replace(g3, "").strip(),
                    }
                ]
            }
            dictionary.append(data)
        elif ws.cell(row=6, column=i).value != None:
            ws.cell(row=4, column=i).value = d1
            ws.cell(row=7, column=i).value = d2
            data = {
                "w": ws.cell(row=1, column=i).value,
                "d": [
                    {
                        "g": replace_all(g1, gros),
                        "e": str(ws.cell(row=4, column=i).value).replace(g1, "").strip(),
                    },
                    {
                        "g": replace_all(g2, gros),
                        "e": str(ws.cell(row=7, column=i).value).replace(g2, "").strip(),
                    }
                ]
            }
            dictionary.append(data)
        else:
            ws.cell(row=4, column=i).value = d1
            data = {
                "w": ws.cell(row=1, column=i).value,
                "d": [
                    {
                        "g": replace_all(g1, gros),
                        "e": str(ws.cell(row=4, column=i).value).replace(g1, "").strip(),
                    }
                ]
            }
            dictionary.append(data)
    filename = str(lws[x])+'.json'
    with open('dictionary.json', 'w', encoding='utf-8') as f:
        json.dump(dictionary, f, ensure_ascii=False, indent=4)
    with open('logs.json', 'w', encoding='utf-8') as g:
        json.dump(logs, g, ensure_ascii=False, indent=4)
    wb.save('results.xlsx')

def replace_all(text, dic):
    for i, j in dic.items():
        text = text.replace(i, j)
    return text


list = range(len(lws))

for i in reversed(list):
    transform(i)
    print("Complete: ", i, ", ", lws[i])
print("DONEZO")

# transform(11)

print(list)
